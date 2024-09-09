import openpyxl
df = openpyxl.load_workbook("Media.xlsx")
film_df = df["Film"]
series_df = df["Series"]
doc_df = df["Documentary"]
clip_df = df["Clip"]


class Database:
    @staticmethod
    def read_from_database():
        # Entering the info of Films into a dictionary
        film_hl = []
        for col in film_df.iter_cols(min_row=1):
            film_hl.append(col[0].value)

        film_list = {key: [] for key in film_hl}

        for row in film_df.iter_rows(min_row=2):
            for n, name in enumerate(film_hl):
                film_list[name].append(row[n].value)

        # Entering the info of Series into a dictionary
        series_hl = []
        for col in series_df.iter_cols(min_row=1):
            series_hl.append(col[0].value)

        series_list = {key: [] for key in series_hl}

        for row in series_df.iter_rows(min_row=2):
            for n, name in enumerate(series_hl):
                series_list[name].append(row[n].value)

        # Entering the info of Documentaries into a dictionary
        doc_hl = []
        for col in doc_df.iter_cols(min_row=1):
            doc_hl.append(col[0].value)

        doc_list = {key: [] for key in doc_hl}

        for row in doc_df.iter_rows(min_row=2):
            for n, name in enumerate(doc_hl):
                doc_list[name].append(row[n].value)

        # Entering the info of Clips into a dictionary
        clip_hl = []
        for col in clip_df.iter_cols(min_row=1):
            clip_hl.append(col[0].value)

        clip_list = {key: [] for key in clip_hl}

        for row in clip_df.iter_rows(min_row=2):
            for n, name in enumerate(clip_hl):
                clip_list[name].append(row[n].value)

        return film_list, series_list, doc_list, clip_list

    @staticmethod
    def write_to_database():
        film_list, series_list, doc_list, clip_list = Database.read_from_database()
        wb = openpyxl.load_workbook("Media.xlsx")
        film_ws = wb["Film"]
        series_ws = wb["Series"]
        doc_ws = wb["Documentary"]
        clip_ws = wb["Clip"]

        # Films
        film_headers = list(film_list.keys())
        for col_num, header in enumerate(film_headers, 1):
            film_ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(zip(*film_list.values()), 2):
            for col_num, value in enumerate(row_data, 1):
                film_ws.cell(row=row_num, column=col_num, value=value)

        # Series
        series_headers = list(series_list.keys())
        for col_num, header in enumerate(series_headers, 1):
            series_ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(zip(*series_list.values()), 2):
            for col_num, value in enumerate(row_data, 1):
                series_ws.cell(row=row_num, column=col_num, value=value)

        # Documentaries
        doc_headers = list(doc_list.keys())
        for col_num, header in enumerate(doc_headers, 1):
            doc_ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(zip(*doc_list.values()), 2):
            for col_num, value in enumerate(row_data, 1):
                doc_ws.cell(row=row_num, column=col_num, value=value)

        # Clips
        clip_headers = list(clip_list.keys())
        for col_num, header in enumerate(clip_headers, 1):
            clip_ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(zip(*doc_list.values()), 2):
            for col_num, value in enumerate(row_data, 1):
                doc_ws.cell(row=row_num, column=col_num, value=value)

        wb.save("Media.xlsx")
